from openpyxl import load_workbook, Workbook

def update_xlsx(file_name, new_row=None):
    try:
        try:
            wb = load_workbook(file_name)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            print(f"{file_name} not found. Creating a new file.")

        if new_row:
            ws.append(new_row)
            print(f"Added new row: {new_row}")

        wb.save(file_name)

    except Exception as e:
        print(f"An error occurred: {e}")

