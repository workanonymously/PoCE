

from __future__ import annotations

import csv
import hashlib
import os
import re
import shutil
import signal
import subprocess
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

try:
    from Combined_frameworks.interact_with_ollama import chat_with_ollama
except ImportError:
    print("[!] Could not import chat_with_ollama from interact_with_ollama.py")
    print("    Put interact_with_ollama.py in the same folder as this script.")
    raise


# =============================================================================
# CONFIG: edit this section only
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent

INPUT_CSV = BASE_DIR / "yasm_poc_data_tagged.csv"

# Use "yasm" if it is available in your PATH.
# Recommended: set this to your ASAN-built YASM binary.
# Example:
# YASM_BIN = "/home/tanu1167/bigstorage/yasm_builds/9defefa/yasm"
YASM_BIN = "yasm"

OUTPUT_DIR = BASE_DIR / "yasm_llm_variant_runs"
ASM_DIR = OUTPUT_DIR / "asm"
LOG_DIR = OUTPUT_DIR / "logs"
OBJ_DIR = OUTPUT_DIR / "objects"
EXCEL_FILE = OUTPUT_DIR / "yasm_llm_crash_results.xlsx"
LLM_RESPONSE_LOG = OUTPUT_DIR / "llm_responses.txt"

# CSV processing
START_ROW = 1       # 1-based row number excluding CSV header
LIMIT = 0           # 0 means process all rows

# Variant generation
RUN_ORIGINAL_POC = True
LLM_VARIANTS_PER_POC = 3

# Ollama
OLLAMA_MODEL = "llama3:70b"
LLM_ATTEMPTS_PER_VARIANT = 3
SLEEP_BETWEEN_LLM_CALLS = 1.0

# YASM execution
TIMEOUT_SEC = 45
YASM_FORMAT = "elf64"  # set to "" if you do not want to pass -f elf64
ASAN_OPTIONS_FIRST = "abort_on_error=1:detect_leaks=0"
ASAN_OPTIONS_SECOND = "abort_on_error=1:detect_leaks=1"
LD_LIBRARY_PATH_PREFIX = ""

REQUIRED_COLUMNS = ["Vulnerable Code", "PoC Exploit Code", "Vul_Index"]


# =============================================================================
# Crash classification
# =============================================================================

SIGNAL_DESC = {
    1: "SIGHUP - terminal line hangup",
    2: "SIGINT - interrupt",
    3: "SIGQUIT - quit",
    4: "SIGILL - illegal instruction",
    5: "SIGTRAP - trace/breakpoint trap",
    6: "SIGABRT - abort",
    7: "SIGBUS - bus error",
    8: "SIGFPE - arithmetic fault",
    9: "SIGKILL - killed",
    11: "SIGSEGV - invalid memory access",
    15: "SIGTERM - termination",
    31: "SIGSYS - bad system call",
}

SUCCESS_KEYWORDS = [
    "SIGSEGV",
    "SIGABRT",
    "SIGBUS",
    "SIGFPE",
    "SIGILL",
    "AddressSanitizer",
    "LeakSanitizer",
    "heap-buffer-overflow",
    "stack-buffer-overflow",
    "global-buffer-overflow",
    "use-after-free",
    "stack-overflow",
    "double free",
    "core dumped",
    "SEGV on unknown address",
    "ABRT",
    "runtime error",
]

HANDLED_ERROR_KEYWORDS = [
    "invalid instruction",
    "not defining a macro",
    "expects a macro identifier",
    "instruction expected",
    "unrecognized character",
    "redefinition of",
    "label or instruction expected",
    "not in a macro call",
    "previously defined here",
    "invalid macro usage",
    "unknown symbol",
    "invalid operand",
    "unterminated macro",
    "expected parameter",
    "instruction expected after label",
    "error: ",
    "warning: ",
]


@dataclass
class RunResult:
    row_number: int
    vul_index: str
    vul_group: str
    variant_number: int
    variant_source: str
    status: str
    crashed: bool
    crash_kind: str
    handled_error: bool
    return_code_first: Optional[int]
    return_code_second: Optional[int]
    signal_number: Optional[int]
    signal_description: str
    elapsed_sec_total: float
    timeout_sec: int
    asm_file: str
    log_file: str
    object_file: str
    code_sha256: str
    code_lines: int
    code_bytes: int
    command_first: str
    command_second: str
    asan_options_first: str
    asan_options_second: str
    vulnerable_code_hash: str
    original_poc_hash: str
    log_excerpt: str


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()


def slugify(value: str, fallback: str = "item") -> str:
    value = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "").strip())
    value = re.sub(r"_+", "_", value).strip("._-")
    return value[:120] or fallback


def short_group(vul_index: str) -> str:
    if not vul_index:
        return "unknown"
    return str(vul_index).split("_")[0]


def normalize_code(raw: str) -> str:
    """Extract clean assembly text from raw CSV text or fenced code."""
    if raw is None:
        return ""

    text = str(raw).replace("\r\n", "\n").replace("\r", "\n").strip()

    fenced = re.findall(
        r"```(?:asm|nasm)?\s*(.*?)```",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if fenced:
        text = fenced[0].strip()

    if text.count("###") >= 2:
        parts = text.split("###")
        if len(parts) >= 3 and parts[1].strip():
            text = parts[1].strip()

    return text.strip() + "\n"


def extract_asm_code(llm_text: str) -> str:
    """Extract one assembly program from the LLM response."""
    if not llm_text:
        return ""

    text = llm_text.replace("\r\n", "\n").replace("\r", "\n").strip()

    fenced = re.findall(
        r"```(?:asm|nasm)?\s*(.*?)```",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if fenced:
        return normalize_code(fenced[0])

    generic = re.findall(r"```\s*(.*?)```", text, flags=re.DOTALL)
    if generic:
        return normalize_code(generic[0])

    if text.count("###") >= 2:
        parts = text.split("###")
        if len(parts) >= 3 and parts[1].strip():
            return normalize_code(parts[1])

    return normalize_code(text)


def looks_like_asm(code: str) -> bool:
    if not code or not code.strip():
        return False

    low = code.lower()
    markers = [
        "%macro",
        "%endmacro",
        "%rep",
        "%endrep",
        "%assign",
        "%define",
        "%undef",
        "%push",
        "%pop",
        "%rotate",
        "%if",
        "%endif",
        "section ",
        "struc",
        "endstruc",
        "istruc",
        "at ",
        "iend",
        "mov ",
        "jmp ",
        "call ",
        "db ",
        "dw ",
        "dd ",
        "resb",
        "times ",
    ]
    return any(marker in low for marker in markers)


def call_ollama(prompt: str) -> str:
    """
    Supports both possible function signatures:
      chat_with_ollama(prompt)
      chat_with_ollama(prompt, model="...")
    """
    try:
        return chat_with_ollama(prompt, model=OLLAMA_MODEL)
    except TypeError:
        return chat_with_ollama(prompt)


def append_llm_log(vul_index: str, variant_number: int, prompt: str, response: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with LLM_RESPONSE_LOG.open("a", encoding="utf-8", errors="replace") as f:
        f.write("\n" + "=" * 100 + "\n")
        f.write(f"Time: {datetime.now().isoformat(timespec='seconds')}\n")
        f.write(f"Vul_Index: {vul_index}\n")
        f.write(f"Variant Number: {variant_number}\n")
        f.write("\n--- PROMPT ---\n")
        f.write(prompt)
        f.write("\n--- RESPONSE ---\n")
        f.write(response or "")
        f.write("\n")


def build_variant_prompt(
    vulnerable_code: str,
    original_poc: str,
    vul_index: str,
    variant_number: int,
) -> str:
    return f"""
You are generating authorized YASM/NASM proof-of-concept test variants
for a controlled local crash-reproduction experiment.

The goal is to generate a new assembly input that is similar in purpose to
the original PoC but not an exact copy. The generated input will be executed
locally against YASM with ASAN enabled.

Vulnerability Index:
{vul_index}

Variant Number:
{variant_number}

Vulnerable Code:
{vulnerable_code}

Original PoC Exploit Code:
{original_poc}

Generate ONE new YASM/NASM assembly PoC variant.

Requirements:
- Output exactly one asm code block.
- Do not include explanation outside the code block.
- Do not include shell commands.
- Do not include C/C++ code.
- Keep it standalone YASM/NASM input.
- Prefer preprocessor or macro-based variation if the original uses macros.
- Do not simply copy the original PoC.
- Avoid global, extern, _start, main, or OS runtime logic unless already present.
- Keep it reasonably small.
- Valid YASM/NASM preprocessor constructs are allowed:
  %macro, %endmacro, %rep, %endrep, %rotate, %define, %assign,
  %if, %endif, %push, %pop, struc, endstruc, istruc, iend,
  token concatenation, scoped macro identifiers, and malformed expressions.

Return only:

```asm
; assembly code here
```
""".strip()


def generate_one_llm_variant(
    vulnerable_code: str,
    original_poc: str,
    vul_index: str,
    variant_number: int,
    existing_hashes: Set[str],
) -> Optional[str]:
    prompt = build_variant_prompt(
        vulnerable_code=vulnerable_code,
        original_poc=original_poc,
        vul_index=vul_index,
        variant_number=variant_number,
    )

    for attempt in range(1, LLM_ATTEMPTS_PER_VARIANT + 1):
        print(
            f"[+] Ollama: generating {vul_index} LLM variant {variant_number} "
            f"(attempt {attempt}/{LLM_ATTEMPTS_PER_VARIANT})"
        )

        response = call_ollama(prompt)
        append_llm_log(vul_index, variant_number, prompt, response)

        code = extract_asm_code(response)

        if not looks_like_asm(code):
            print(f"[!] Rejected LLM output: does not look like asm ({vul_index}, variant {variant_number})")
            time.sleep(SLEEP_BETWEEN_LLM_CALLS)
            continue

        code_hash = sha256_text(code)
        if code_hash in existing_hashes:
            print(f"[!] Rejected LLM output: duplicate code ({vul_index}, variant {variant_number})")
            time.sleep(SLEEP_BETWEEN_LLM_CALLS)
            continue

        existing_hashes.add(code_hash)
        return code

    return None


def generate_variants_for_row(
    vulnerable_code: str,
    original_poc_raw: str,
    vul_index: str,
) -> List[Tuple[str, str]]:
    """
    Returns:
        [(asm_code, variant_source), ...]

    If RUN_ORIGINAL_POC=True, original CSV PoC is included first.
    Then LLM_VARIANTS_PER_POC new variants are generated through Ollama.
    """
    variants: List[Tuple[str, str]] = []

    original_poc = normalize_code(original_poc_raw)
    existing_hashes: Set[str] = set()

    if RUN_ORIGINAL_POC and original_poc.strip():
        variants.append((original_poc, "original_csv"))
        existing_hashes.add(sha256_text(original_poc))

    for variant_number in range(1, LLM_VARIANTS_PER_POC + 1):
        generated = generate_one_llm_variant(
            vulnerable_code=vulnerable_code,
            original_poc=original_poc,
            vul_index=vul_index,
            variant_number=variant_number,
            existing_hashes=existing_hashes,
        )

        if generated:
            variants.append((generated, "llm_generated"))
        else:
            print(f"[!] Failed to generate LLM variant {variant_number} for {vul_index}")

    return variants


def load_csv_rows(csv_path: Path, start_row: int, limit: int) -> List[Dict[str, str]]:
    if not csv_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {csv_path}")

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        missing = [col for col in REQUIRED_COLUMNS if col not in (reader.fieldnames or [])]
        if missing:
            raise ValueError(
                f"Missing required CSV columns: {missing}. "
                f"Found columns: {reader.fieldnames}"
            )
        rows = list(reader)

    start_idx = max(0, int(start_row or 1) - 1)
    rows = rows[start_idx:]

    if limit and int(limit) > 0:
        rows = rows[: int(limit)]

    return rows


def classify_log(
    log: str,
    return_code: Optional[int],
    timed_out: bool,
) -> Tuple[bool, str, bool, Optional[int], str]:
    """
    Returns:
        crashed, crash_kind, handled_error, signal_number, signal_description
    """
    signal_number: Optional[int] = None
    signal_description = ""

    if timed_out:
        return True, "TIMEOUT", False, None, "Timeout"

    if return_code is not None and return_code < 0:
        signal_number = -return_code
        signal_description = SIGNAL_DESC.get(signal_number, f"Signal {signal_number}")

        if signal_number == signal.SIGSEGV:
            return True, "SEGV", False, signal_number, signal_description
        if signal_number == signal.SIGABRT:
            return True, "ABRT", False, signal_number, signal_description
        if signal_number == signal.SIGBUS:
            return True, "BUS", False, signal_number, signal_description
        if signal_number == signal.SIGFPE:
            return True, "FPE", False, signal_number, signal_description
        if signal_number == signal.SIGILL:
            return True, "ILL", False, signal_number, signal_description

        return True, f"SIGNAL_{signal_number}", False, signal_number, signal_description

    low = log.lower()
    handled = any(k.lower() in low for k in HANDLED_ERROR_KEYWORDS)

    if "heap-buffer-overflow" in low:
        return True, "HEAPBOF", handled, signal_number, signal_description
    if "stack-buffer-overflow" in low:
        return True, "STACKBOF", handled, signal_number, signal_description
    if "global-buffer-overflow" in low:
        return True, "GLOBALBOF", handled, signal_number, signal_description
    if "use-after-free" in low:
        return True, "UAF", handled, signal_number, signal_description
    if "stack-overflow" in low:
        return True, "STACK_OVERFLOW", handled, signal_number, signal_description
    if "double free" in low:
        return True, "DOUBLE_FREE", handled, signal_number, signal_description
    if "addresssanitizer" in low:
        return True, "ASAN", handled, signal_number, signal_description
    if "leaksanitizer" in low:
        return True, "LSAN", handled, signal_number, signal_description

    if any(k.lower() in low for k in SUCCESS_KEYWORDS):
        # Avoid treating ordinary parser errors as crashes.
        if handled and not any(k in low for k in ["signal", "asan", "sanitizer", "core dumped"]):
            return False, "HANDLED_ERROR", True, signal_number, signal_description
        return True, "GENCRASH", handled, signal_number, signal_description

    if handled:
        return False, "HANDLED_ERROR", True, signal_number, signal_description

    return False, "NO_CRASH", False, signal_number, signal_description


def run_yasm_once(
    yasm_bin: str,
    asm_path: Path,
    obj_path: Path,
    timeout_sec: int,
    fmt: str,
    asan_options: str,
    ld_library_path_prefix: str,
) -> Tuple[str, Optional[int], float, bool, str]:
    cmd = [yasm_bin]

    if fmt:
        cmd += ["-f", fmt]

    cmd += [str(asm_path), "-o", str(obj_path)]

    env = os.environ.copy()
    env["ASAN_OPTIONS"] = asan_options

    if ld_library_path_prefix:
        old_ld = env.get("LD_LIBRARY_PATH", "")
        env["LD_LIBRARY_PATH"] = ld_library_path_prefix + (":" + old_ld if old_ld else "")

    start = time.time()
    timed_out = False

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        env=env,
        preexec_fn=os.setsid if hasattr(os, "setsid") else None,
    )

    try:
        stdout, stderr = proc.communicate(timeout=timeout_sec)
    except subprocess.TimeoutExpired:
        timed_out = True
        try:
            if hasattr(os, "killpg"):
                os.killpg(os.getpgid(proc.pid), signal.SIGTERM)
            else:
                proc.kill()
        except Exception:
            proc.kill()

        stdout, stderr = proc.communicate()

    elapsed = time.time() - start
    rc = proc.returncode

    full_log = (
        f"Command: {' '.join(cmd)}\n"
        f"Return code: {rc}\n"
        f"Elapsed seconds: {elapsed:.3f}\n"
        f"Timed out: {timed_out}\n"
        f"ASAN_OPTIONS: {asan_options}\n"
        f"\n--- STDOUT ---\n{stdout or ''}\n"
        f"\n--- STDERR ---\n{stderr or ''}\n"
    )

    if rc is not None and rc < 0:
        sig = -rc
        full_log += f"\n[Killed by signal {sig} - {SIGNAL_DESC.get(sig, 'unknown signal')}]\n"

    if timed_out:
        full_log += f"\n[Timeout after {timeout_sec} seconds]\n"

    return full_log, rc, elapsed, timed_out, " ".join(cmd)


def run_yasm_with_two_asan_modes(
    yasm_bin: str,
    asm_path: Path,
    obj_path: Path,
) -> Tuple[str, Optional[int], Optional[int], float, bool, str, str]:
    """
    Run first with leak detection off.
    If no successful crash is detected, run again with leak detection on.
    """
    log1, rc1, elapsed1, timeout1, cmd1 = run_yasm_once(
        yasm_bin=yasm_bin,
        asm_path=asm_path,
        obj_path=obj_path,
        timeout_sec=TIMEOUT_SEC,
        fmt=YASM_FORMAT,
        asan_options=ASAN_OPTIONS_FIRST,
        ld_library_path_prefix=LD_LIBRARY_PATH_PREFIX,
    )

    crashed1, _, _, _, _ = classify_log(log1, rc1, timeout1)

    if crashed1:
        return log1, rc1, None, elapsed1, timeout1, cmd1, ""

    obj_path_second = obj_path.with_suffix(".asan2.o")
    log2, rc2, elapsed2, timeout2, cmd2 = run_yasm_once(
        yasm_bin=yasm_bin,
        asm_path=asm_path,
        obj_path=obj_path_second,
        timeout_sec=TIMEOUT_SEC,
        fmt=YASM_FORMAT,
        asan_options=ASAN_OPTIONS_SECOND,
        ld_library_path_prefix=LD_LIBRARY_PATH_PREFIX,
    )

    combined_log = (
        "========== FIRST RUN: detect_leaks=0 ==========\n"
        + log1
        + "\n\n========== SECOND RUN: detect_leaks=1 ==========\n"
        + log2
    )

    return combined_log, rc1, rc2, elapsed1 + elapsed2, timeout1 or timeout2, cmd1, cmd2


def excel_safe(value):
    if value is None:
        return ""

    if isinstance(value, str):
        value = value.replace("\x00", "")
        if value and value[0] in ("=", "+", "-", "@"):
            return "'" + value

    return value


def write_excel(results: List[RunResult], excel_path: Path, metadata: Dict[str, str]) -> None:
    if Workbook is None:
        raise RuntimeError(
            "openpyxl is required to write .xlsx files. Install it with:\n"
            "pip install openpyxl"
        )

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_results = wb.create_sheet("Results")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    light_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total = len(results)
    crashed = sum(1 for r in results if r.crashed)
    handled = sum(1 for r in results if r.handled_error)
    timeouts = sum(1 for r in results if r.crash_kind == "TIMEOUT")
    unique_vuls = len({r.vul_index for r in results})

    summary_rows = [
        ["YASM LLM Variant Crash Report", ""],
        ["Generated At", metadata.get("generated_at", "")],
        ["Input CSV", metadata.get("input_csv", "")],
        ["YASM Binary", metadata.get("yasm_bin", "")],
        ["Output Directory", metadata.get("output_dir", "")],
        ["Excel File", str(excel_path)],
        ["Run Original PoC", str(RUN_ORIGINAL_POC)],
        ["LLM Variants Per PoC", LLM_VARIANTS_PER_POC],
        ["Ollama Model", OLLAMA_MODEL],
        ["Total Runs", total],
        ["Unique Vul_Index Values", unique_vuls],
        ["Crashed", crashed],
        ["Handled Errors", handled],
        ["Timeouts", timeouts],
        ["Crash Rate", f"{(crashed / total * 100):.2f}%" if total else "0.00%"],
    ]

    for row in summary_rows:
        ws_summary.append(row)

    ws_summary["A1"].font = Font(bold=True, size=14, color="1F4E78")

    for row in ws_summary.iter_rows(
        min_row=2,
        max_row=len(summary_rows),
        min_col=1,
        max_col=2,
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].font = Font(bold=True)
        row[0].fill = light_fill

    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 100

    result_dicts = [asdict(r) for r in results]
    headers = list(result_dicts[0].keys()) if result_dicts else [
        "row_number",
        "vul_index",
        "variant_number",
        "variant_source",
        "status",
        "crashed",
        "crash_kind",
        "asm_file",
        "log_file",
    ]

    ws_results.append(headers)
    for cell in ws_results[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rd in result_dicts:
        ws_results.append([excel_safe(rd.get(h, "")) for h in headers])

    for row in ws_results.iter_rows(
        min_row=2,
        max_row=ws_results.max_row,
        min_col=1,
        max_col=ws_results.max_column,
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    status_col = headers.index("status") + 1 if "status" in headers else None
    crash_col = headers.index("crashed") + 1 if "crashed" in headers else None

    for r in range(2, ws_results.max_row + 1):
        status = str(ws_results.cell(r, status_col).value) if status_col else ""
        crashed_value = ws_results.cell(r, crash_col).value if crash_col else False

        if status == "CRASH" or str(crashed_value).lower() == "true":
            fill = PatternFill("solid", fgColor="E2F0D9")
        elif status == "HANDLED_ERROR":
            fill = PatternFill("solid", fgColor="FFF2CC")
        else:
            fill = PatternFill("solid", fgColor="FCE4D6")

        for c in range(1, ws_results.max_column + 1):
            ws_results.cell(r, c).fill = fill

    ws_results.freeze_panes = "A2"
    ws_results.auto_filter.ref = ws_results.dimensions

    width_caps = {
        "log_excerpt": 90,
        "command_first": 70,
        "command_second": 70,
        "asm_file": 60,
        "log_file": 60,
        "object_file": 60,
        "asan_options_first": 35,
        "asan_options_second": 35,
        "signal_description": 35,
    }

    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)

        if header in width_caps:
            ws_results.column_dimensions[letter].width = width_caps[header]
        elif header in {
            "crashed",
            "handled_error",
            "return_code_first",
            "return_code_second",
            "signal_number",
            "elapsed_sec_total",
            "timeout_sec",
            "code_lines",
            "code_bytes",
        }:
            ws_results.column_dimensions[letter].width = 16
        elif header.endswith("hash") or header == "code_sha256":
            ws_results.column_dimensions[letter].width = 22
        else:
            ws_results.column_dimensions[letter].width = min(max(len(header) + 4, 16), 35)

    for ws in [ws_summary, ws_results]:
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 18

    ws_results.row_dimensions[1].height = 30

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def resolve_yasm_path(yasm_bin: str) -> Tuple[str, bool]:
    if os.path.isabs(yasm_bin):
        return yasm_bin, Path(yasm_bin).exists()

    found = shutil.which(yasm_bin)
    if found:
        return found, True

    return yasm_bin, False


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ASM_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    OBJ_DIR.mkdir(parents=True, exist_ok=True)

    input_csv = INPUT_CSV.expanduser().resolve()
    excel_path = EXCEL_FILE.expanduser().resolve()

    yasm_path, yasm_exists = resolve_yasm_path(YASM_BIN)

    if not yasm_exists:
        print(f"[!] YASM executable not found: {YASM_BIN}", file=sys.stderr)
        print("    Edit YASM_BIN at the top of this script.", file=sys.stderr)
        print("    Example:", file=sys.stderr)
        print('    YASM_BIN = "/home/tanu1167/bigstorage/yasm_builds/9defefa/yasm"', file=sys.stderr)
        return 2

    rows = load_csv_rows(input_csv, START_ROW, LIMIT)

    print(f"[+] Loaded CSV rows: {len(rows)}")
    print(f"[+] Input CSV: {input_csv}")
    print(f"[+] Output directory: {OUTPUT_DIR}")
    print(f"[+] Excel file: {excel_path}")
    print(f"[+] YASM binary: {yasm_path}")
    print(f"[+] RUN_ORIGINAL_POC: {RUN_ORIGINAL_POC}")
    print(f"[+] LLM_VARIANTS_PER_POC: {LLM_VARIANTS_PER_POC}")
    print(f"[+] Ollama model: {OLLAMA_MODEL}")

    results: List[RunResult] = []
    run_counter = 0

    for csv_idx, row in enumerate(rows, start=START_ROW):
        vul_index = str(row.get("Vul_Index", "")).strip() or f"row_{csv_idx}"
        vul_group = short_group(vul_index)
        vulnerable_code = str(row.get("Vulnerable Code", "") or "")
        original_poc_raw = str(row.get("PoC Exploit Code", "") or "")
        original_poc_norm = normalize_code(original_poc_raw)

        print("\n" + "=" * 90)
        print(f"[+] Processing CSV row {csv_idx}, Vul_Index={vul_index}")
        print("=" * 90)

        variants = generate_variants_for_row(
            vulnerable_code=vulnerable_code,
            original_poc_raw=original_poc_raw,
            vul_index=vul_index,
        )

        if not variants:
            print(f"[!] No variants generated for {vul_index}; skipping.")
            continue

        for variant_idx, (variant_code, variant_source) in enumerate(variants, start=1):
            run_counter += 1

            base_name = (
                f"{csv_idx:04d}_{slugify(vul_index)}_"
                f"v{variant_idx}_{slugify(variant_source)}"
            )

            asm_path = ASM_DIR / f"{base_name}.asm"
            log_path = LOG_DIR / f"{base_name}.log"
            obj_path = OBJ_DIR / f"{base_name}.o"

            asm_path.write_text(variant_code, encoding="utf-8", errors="replace")

            code_hash = sha256_text(variant_code)
            code_lines = len(variant_code.splitlines())
            code_bytes = len(variant_code.encode("utf-8", errors="replace"))
            vuln_hash = sha256_text(vulnerable_code) if vulnerable_code.strip() else ""
            original_hash = sha256_text(original_poc_norm) if original_poc_norm.strip() else ""

            full_log, rc1, rc2, elapsed_total, timed_out, cmd1, cmd2 = run_yasm_with_two_asan_modes(
                yasm_bin=str(yasm_path),
                asm_path=asm_path,
                obj_path=obj_path,
            )

            crashed, crash_kind, handled_error, sig_num, sig_desc = classify_log(
                log=full_log,
                return_code=rc2 if rc2 is not None else rc1,
                timed_out=timed_out,
            )

            if crashed:
                status = "CRASH"
            elif handled_error:
                status = "HANDLED_ERROR"
            else:
                status = "NO_CRASH"

            log_path.write_text(full_log, encoding="utf-8", errors="replace")
            excerpt = full_log[-6000:] if len(full_log) > 6000 else full_log

            results.append(
                RunResult(
                    row_number=csv_idx,
                    vul_index=vul_index,
                    vul_group=vul_group,
                    variant_number=variant_idx,
                    variant_source=variant_source,
                    status=status,
                    crashed=crashed,
                    crash_kind=crash_kind,
                    handled_error=handled_error,
                    return_code_first=rc1,
                    return_code_second=rc2,
                    signal_number=sig_num,
                    signal_description=sig_desc,
                    elapsed_sec_total=round(elapsed_total, 3),
                    timeout_sec=TIMEOUT_SEC,
                    asm_file=str(asm_path),
                    log_file=str(log_path),
                    object_file=str(obj_path),
                    code_sha256=code_hash,
                    code_lines=code_lines,
                    code_bytes=code_bytes,
                    command_first=cmd1,
                    command_second=cmd2,
                    asan_options_first=ASAN_OPTIONS_FIRST,
                    asan_options_second=ASAN_OPTIONS_SECOND,
                    vulnerable_code_hash=vuln_hash,
                    original_poc_hash=original_hash,
                    log_excerpt=excerpt,
                )
            )

            print(
                f"[{run_counter}] {vul_index} | variant {variant_idx} | "
                f"{variant_source} | {status} ({crash_kind})"
            )

    metadata = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_csv": str(input_csv),
        "yasm_bin": str(yasm_path),
        "output_dir": str(OUTPUT_DIR),
    }

    write_excel(results, excel_path, metadata)

    print("\n[+] Complete")
    print(f"[+] Excel report: {excel_path}")
    print(f"[+] ASM files:     {ASM_DIR}")
    print(f"[+] Logs:          {LOG_DIR}")
    print(f"[+] Object files:  {OBJ_DIR}")
    print(f"[+] LLM log:       {LLM_RESPONSE_LOG}")

    crashed_count = sum(1 for r in results if r.crashed)
    print(f"[+] Crashes:       {crashed_count}/{len(results)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

